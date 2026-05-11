# excel_formatting.py

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# ======================================================
# MAIN ENTRY
# ======================================================

def format_report_workbook(wb):
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        apply_default_border(ws)
        apply_default_alignment(ws)
        auto_width(ws)

    if "Weekly Timesheet" in wb.sheetnames:
        format_weekly_or_roster_sheet(wb["Weekly Timesheet"])

    if "Roster Summary" in wb.sheetnames:
        format_weekly_or_roster_sheet(wb["Roster Summary"])

    if "Gap Summary" in wb.sheetnames:
        format_gap_summary_sheet(wb["Gap Summary"])


# ======================================================
# COMMON STYLE
# ======================================================

def apply_default_border(ws):
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border


def apply_default_alignment(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                vertical="center",
                horizontal="center",
                wrap_text=True
            )


def auto_width(ws):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0

        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))

        width = min(max_len + 3, 38)

        # Narrow common day/hour columns
        if max_len <= 5:
            width = 10

        ws.column_dimensions[col_letter].width = width


# ======================================================
# WEEKLY TIMESHEET / ROSTER SUMMARY
# ======================================================

def format_weekly_or_roster_sheet(ws):
    ws.freeze_panes = "A8"

    dark_blue = PatternFill("solid", fgColor="17365D")
    mid_blue = PatternFill("solid", fgColor="1F4E78")
    light_blue = PatternFill("solid", fgColor="D9EAF7")
    total_yellow = PatternFill("solid", fgColor="FFF2CC")
    grey_fill = PatternFill("solid", fgColor="F2F2F2")

    white_bold = Font(color="FFFFFF", bold=True)
    black_bold = Font(color="000000", bold=True)

    # Row height
    for row_num in range(1, ws.max_row + 1):
        ws.row_dimensions[row_num].height = 24

    # Top 8 rows are report header / grouped header area
    for row_num in range(1, min(ws.max_row, 8) + 1):
        for cell in ws[row_num]:
            if cell.value is None:
                continue

            if row_num == 1:
                cell.fill = dark_blue
                cell.font = Font(color="FFFFFF", bold=True, size=13)
            elif row_num in [2, 3, 4]:
                cell.fill = light_blue
                cell.font = black_bold
            else:
                cell.fill = mid_blue
                cell.font = white_bold

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

    # Body formatting
    for row in ws.iter_rows(min_row=9):
        for cell in row:
            value = str(cell.value).strip().lower() if cell.value is not None else ""

            # Total row / total cells
            if "total" in value:
                for c in ws[cell.row]:
                    c.fill = total_yellow
                    c.font = black_bold

            # Light grey for empty-looking separator rows
            if value == "":
                cell.fill = grey_fill

            # Number format
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"

    # Make first few columns wider
    fixed_widths = {
        "A": 18,
        "B": 22,
        "C": 18,
        "D": 18,
        "E": 26,
        "F": 26,
        "G": 18,
        "H": 18,
    }

    for col, width in fixed_widths.items():
        ws.column_dimensions[col].width = width


# ======================================================
# GAP SUMMARY
# ======================================================

def format_gap_summary_sheet(ws):
    ws.freeze_panes = "A2"

    dark_blue = PatternFill("solid", fgColor="1F4E78")
    dark_red = PatternFill("solid", fgColor="C00000")
    light_red = PatternFill("solid", fgColor="FDE9E7")
    white_bold = Font(color="FFFFFF", bold=True)
    red_font = Font(color="9C0006", bold=True)

    # 找右边 unmatched 区域：
    # 当前右边表头是 OPMS | SignOutTime | NextSignInTime | HoursBetween | Reason
    # 所以用 SignOutTime / NextSignInTime 定位，再往左一列回到 OPMS
    unmatched_start_col = None

    for cell in ws[1]:
        value = str(cell.value).strip().lower() if cell.value else ""

        if value in ["signouttime", "nextsignintime"]:
            unmatched_start_col = cell.column - 1
            break

    # 左边 Gap Summary 结束列
    main_end_col = ws.max_column

    if unmatched_start_col:
        main_end_col = unmatched_start_col - 2

    # 左边主表 header 蓝色
    for col in range(1, main_end_col + 1):
        cell = ws.cell(row=1, column=col)

        if cell.value is not None:
            cell.fill = dark_blue
            cell.font = white_bold
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

    # 中间空列保持空白，不要 Column_4
    if unmatched_start_col and unmatched_start_col > 1:
        spacer_col = unmatched_start_col - 1
        for row_num in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_num, column=spacer_col)
            cell.fill = PatternFill(fill_type=None)
            cell.value = None

    # 右边 unmatched header 红色
    if unmatched_start_col:
        for col in range(unmatched_start_col, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)

            if cell.value is not None:
                cell.fill = dark_red
                cell.font = white_bold
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )

        # 右边 unmatched body 浅红
        for row_num in range(2, ws.max_row + 1):
            for col in range(unmatched_start_col, ws.max_column + 1):
                cell = ws.cell(row=row_num, column=col)

                if cell.value is not None:
                    cell.fill = light_red
                    cell.font = red_font

    # 左边 unresolved / missing 行标红
    for row in ws.iter_rows(min_row=2, max_col=main_end_col):
        row_has_problem = False

        for cell in row:
            value = str(cell.value).lower() if cell.value is not None else ""

            if (
                "unresolved" in value
                or "missing" in value
                or "no sign in" in value
                or "no sign out" in value
            ):
                row_has_problem = True

        if row_has_problem:
            for cell in row:
                cell.fill = light_red
                cell.font = red_font

    # 数字格式
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"

    auto_width(ws)