import os
import requests
from datetime import datetime, date, timedelta
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from TimesheethourFormat import format_report_workbook
from io import BytesIO

load_dotenv()

# ======================================================
# ENV CONFIG
# ======================================================

TENANT_ID = os.getenv("SHAREPOINT_TENANT_ID")
CLIENT_ID = os.getenv("SHAREPOINT_CLIENT_ID")
CLIENT_SECRET = os.getenv("SHAREPOINT_CLIENT_SECRET")

SHAREPOINT_HOST = os.getenv("SHAREPOINT_HOST")
SITE_NAME = os.getenv("SITE_NAME")

ROSTER_LIST_NAME = os.getenv("LIST_NAME", "PPL-Rosters")
TIMESHEET_LIST_NAME = os.getenv("GAP_LIST_NAME", "PPL-Timesheets")

SUPPLIER_LIST_NAME = os.getenv("SUPPLIER_LIST_NAME", "SMS-Suppliers")
SITE_LOOKUP_LIST_NAME = os.getenv("SITE_LOOKUP_LIST_NAME", "SYS-OpsSections")

MAX_CROSS_DAY_GAP_HOURS = 6


# ======================================================
# AUTH
# ======================================================

def get_access_token():

    url = (
        f"https://login.microsoftonline.com/"
        f"{TENANT_ID}/oauth2/v2.0/token"
    )

    payload = {
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "scope": "https://graph.microsoft.com/.default",
        "grant_type": "client_credentials",
    }

    response = requests.post(url, data=payload)

    if not response.ok:
        print("❌ TOKEN ERROR")
        print(response.text)

    response.raise_for_status()

    return response.json()["access_token"]


# ======================================================
# SHAREPOINT HELPERS
# ======================================================

def get_site_id(token):

    url = (
        f"https://graph.microsoft.com/v1.0/sites/"
        f"{SHAREPOINT_HOST}:/sites/{SITE_NAME}"
    )

    headers = {
        "Authorization": f"Bearer {token}"
    }

    response = requests.get(url, headers=headers)

    if not response.ok:
        print("❌ SITE ERROR")
        print(response.text)

    response.raise_for_status()

    site_id = response.json()["id"]

    print(f"✅ Site ID: {site_id}")

    return site_id


def get_list_id(token, site_id, list_name):

    url = (
        f"https://graph.microsoft.com/v1.0/sites/"
        f"{site_id}/lists"
    )

    headers = {
        "Authorization": f"Bearer {token}"
    }

    response = requests.get(url, headers=headers)

    if not response.ok:
        print("❌ LIST FETCH ERROR")
        print(response.text)

    response.raise_for_status()

    target = list_name.strip().lower()

    for item in response.json().get("value", []):

        display_name = item.get("displayName", "").strip()
        internal_name = item.get("name", "").strip()

        print(
            f"Checking list -> "
            f"displayName='{display_name}', "
            f"name='{internal_name}'"
        )

        if (
            display_name.lower() == target
            or internal_name.lower() == target
        ):
            print(f"✅ Found list: {display_name}")
            return item["id"]

    raise Exception(f"List not found: {list_name}")


def fetch_all_list_items(token, site_id, list_id):

    headers = {
        "Authorization": f"Bearer {token}"
    }

    all_items = []

    url = (
        f"https://graph.microsoft.com/v1.0/sites/"
        f"{site_id}/lists/{list_id}/items"
        f"?$expand=fields&$top=5000"
    )

    while url:

        response = requests.get(url, headers=headers)

        if not response.ok:
            print("❌ FETCH ITEMS ERROR")
            print(response.text)

        response.raise_for_status()

        data = response.json()

        batch = data.get("value", [])

        all_items.extend(batch)

        print(f"✅ Pulled {len(batch)} rows")

        url = data.get("@odata.nextLink")

    print(f"✅ Total rows: {len(all_items)}")

    return all_items


# ======================================================
# DATE HELPERS
# ======================================================

def parse_sharepoint_date(value):

    if not value:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    value = str(value).strip()

    if value.endswith("Z") and "T" in value:

        dt_utc = datetime.strptime(
            value,
            "%Y-%m-%dT%H:%M:%SZ"
        )

        dt_perth = dt_utc + timedelta(hours=8)

        return dt_perth.date()

    formats = [
        "%d/%m/%Y",
        "%d/%m/%Y %I:%M %p",
        "%d/%m/%Y %H:%M",
        "%Y-%m-%dT%H:%M:%S",
    ]

    for fmt in formats:
        try:
            return datetime.strptime(value, fmt).date()
        except:
            pass

    raise ValueError(f"Unsupported date format: {value}")


def parse_sharepoint_datetime(value):

    if not value:
        return None

    if isinstance(value, datetime):
        return value

    value = str(value).strip()

    if value.endswith("Z") and "T" in value:

        dt_utc = datetime.strptime(
            value,
            "%Y-%m-%dT%H:%M:%SZ"
        )

        dt_perth = dt_utc + timedelta(hours=8)

        return dt_perth

    formats = [
        "%d/%m/%Y %I:%M %p",
        "%d/%m/%Y %H:%M",
        "%Y-%m-%dT%H:%M:%S",
    ]

    for fmt in formats:
        try:
            return datetime.strptime(value, fmt)
        except:
            pass

    raise ValueError(f"Unsupported datetime format: {value}")


def get_last_week_range():

    today = datetime.today().date()

    this_monday = today - timedelta(days=today.weekday())

    last_monday = this_monday - timedelta(days=7)
    last_sunday = this_monday - timedelta(days=1)

    return last_monday, last_sunday


# ======================================================
# LOOKUP MAPS
# ======================================================

def build_lookup_map(lookup_items):

    result = {}

    for item in lookup_items:

        item_id = str(item.get("id", "")).strip()

        fields = item.get("fields", {})

        title = str(fields.get("Title", "")).strip()

        if item_id and title:
            result[item_id] = title

    return result


def get_lookup_id(fields, base_name):

    possible_fields = [
        f"{base_name}LookupId",
        f"{base_name}Id",
        base_name,
    ]

    for field_name in possible_fields:

        value = fields.get(field_name)

        if value is None:
            continue

        if isinstance(value, dict):
            value = value.get("Id") or value.get("id")

        value_str = str(value).strip()

        if value_str:
            return value_str

    return ""


def get_supplier_value(fields, supplier_map):

    direct_value = (
        fields.get("SupplierLookupValue")
        or fields.get("Supplier_x003a_Title")
    )

    if direct_value and str(direct_value).strip():
        return str(direct_value).strip()

    supplier_id = get_lookup_id(fields, "Supplier")

    if not supplier_id:
        return ""

    return supplier_map.get(supplier_id, "")


def get_site_value(fields, site_map):

    direct_value = (
        fields.get("SiteLookupValue")
        or fields.get("Site_x003a_Title")
    )

    if direct_value and str(direct_value).strip():
        return str(direct_value).strip()

    site_id = get_lookup_id(fields, "Site")

    if not site_id:
        return ""

    return site_map.get(site_id, "")


# ======================================================
# BUSINESS LOGIC
# ======================================================

def is_asset_transport_hire(position, site_name):

    position_text = str(position or "").strip().upper()
    site_text = str(site_name or "").strip().upper()

    return (
        position_text.startswith("Z.")
        and site_text == "TRANSPORT & HIRE"
    )


def get_shift_type(work_type):

    wt = str(work_type or "").strip().upper()

    if wt == "NIGHT SHIFT":
        return "NS"

    return "DS"


def get_project_value(fields):

    return (
        fields.get("ProjectLookupValue")
        or fields.get("Project")
        or fields.get("ClientName")
        or ""
    )


def get_roster_hours(fields):

    possible_hour_fields = [
        "Hours",
        "CurrentHours",
        "Current_x0020_Hours",
        "RosterHours",
        "Roster_x0020_Hours",
    ]

    for field_name in possible_hour_fields:

        value = fields.get(field_name)

        if value is None:
            continue

        value_str = str(value).strip()

        if value_str == "":
            continue

        try:
            return float(value_str)

        except:
            print(
                f"⚠️ Invalid Hours value: "
                f"{field_name}={value}"
            )

            return 0.0

    return 0.0


# ======================================================
# GAP CALCULATION
# ======================================================

def build_gap_map(
    timesheet_items,
    week_start,
    week_end
):

    records_by_opms = {}

    unmatched_signout_rows = []

    for item in timesheet_items:

        fields = item.get("fields", {})

        opms = str(fields.get("OPMS", "")).strip()

        status = str(
            fields.get("Status", "")
        ).strip().lower()

        dt_raw = fields.get("Date")

        if not opms or not status or not dt_raw:
            continue

        dt = parse_sharepoint_datetime(dt_raw)

        if not dt:
            continue

        work_date = dt.date()

        # 多取一天，给 night shift 跨天 Sign In 用
        if (
            work_date < week_start
            or work_date > week_end + timedelta(days=1)
        ):
            continue

        records_by_opms.setdefault(opms, []).append({
            "datetime": dt,
            "status": status,
        })

    gap_map = {}

    for opms, records in records_by_opms.items():

        records.sort(
            key=lambda x: x["datetime"]
        )

        open_sign_out = None

        for record in records:

            status = record["status"]
            current_time = record["datetime"]

            # ==================================================
            # SIGN OUT
            # ==================================================

            if status == "sign out":

                if open_sign_out:

                    unmatched_signout_rows.append({
                        "OPMS": opms,
                        "SignOutTime": open_sign_out,
                        "NextSignInTime": "",
                        "HoursBetween": "",
                        "Reason": "No matching Sign In found",
                    })

                open_sign_out = current_time

            # ==================================================
            # SIGN IN
            # ==================================================

            elif status == "sign in":

                if not open_sign_out:
                    continue

                start_dt = open_sign_out
                end_dt = current_time

                total_diff_hours = (
                    end_dt - start_dt
                ).total_seconds() / 3600

                # Sign In 必须大于 Sign Out
                if total_diff_hours <= 0:

                    unmatched_signout_rows.append({
                        "OPMS": opms,
                        "SignOutTime": start_dt,
                        "NextSignInTime": end_dt,
                        "HoursBetween": round(total_diff_hours, 2),
                        "Reason": "Ignored - Sign In <= Sign Out",
                    })

                    open_sign_out = None
                    continue

                # 跨天不能无限配对
                if (
                    start_dt.date() != end_dt.date()
                    and total_diff_hours > MAX_CROSS_DAY_GAP_HOURS
                ):

                    unmatched_signout_rows.append({
                        "OPMS": opms,
                        "SignOutTime": start_dt,
                        "NextSignInTime": end_dt,
                        "HoursBetween": round(total_diff_hours, 2),
                        "Reason": (
                            "Ignored - cross-day gap over "
                            f"{MAX_CROSS_DAY_GAP_HOURS} hours"
                        ),
                    })

                    open_sign_out = None
                    continue

                # 合法 gap：按天拆分
                current_day = start_dt.date()
                end_day = end_dt.date()

                while current_day <= end_day:

                    day_start = datetime.combine(
                        current_day,
                        datetime.min.time()
                    )

                    day_end = datetime.combine(
                        current_day + timedelta(days=1),
                        datetime.min.time()
                    )

                    split_start = max(
                        start_dt,
                        day_start
                    )

                    split_end = min(
                        end_dt,
                        day_end
                    )

                    diff_hours = (
                        split_end - split_start
                    ).total_seconds() / 3600

                    if diff_hours > 0:

                        key = (
                            opms,
                            current_day
                        )

                        gap_map[key] = (
                            gap_map.get(key, 0.0)
                            + diff_hours
                        )

                    current_day += timedelta(days=1)

                open_sign_out = None

        # 最后一条 Sign Out 没有 Sign In
        if open_sign_out:

            unmatched_signout_rows.append({
                "OPMS": opms,
                "SignOutTime": open_sign_out,
                "NextSignInTime": "",
                "HoursBetween": "",
                "Reason": "No matching Sign In found",
            })

    for key, value in list(gap_map.items()):
        gap_map[key] = round(value, 2)

    gap_summary_rows = []

    for key, value in gap_map.items():

        opms, work_date = key

        if week_start <= work_date <= week_end:

            gap_summary_rows.append({
                "OPMS": opms,
                "Date": work_date,
                "TotalGapHours": round(value, 2),
            })

    gap_summary_rows.sort(
        key=lambda x: (
            x["OPMS"],
            x["Date"]
        )
    )

    unmatched_signout_rows.sort(
        key=lambda x: (
            x["OPMS"],
            str(x["SignOutTime"])
        )
    )

    return (
        gap_map,
        gap_summary_rows,
        unmatched_signout_rows
    )


# ======================================================
# DAILY RESULTS
# ======================================================

def calculate_daily_results():

    week_start, week_end = get_last_week_range()

    token = get_access_token()

    site_id = get_site_id(token)

    roster_list_id = get_list_id(
        token,
        site_id,
        ROSTER_LIST_NAME
    )

    timesheet_list_id = get_list_id(
        token,
        site_id,
        TIMESHEET_LIST_NAME
    )

    supplier_list_id = get_list_id(
        token,
        site_id,
        SUPPLIER_LIST_NAME
    )

    site_lookup_list_id = get_list_id(
        token,
        site_id,
        SITE_LOOKUP_LIST_NAME
    )

    roster_items = fetch_all_list_items(
        token,
        site_id,
        roster_list_id
    )

    timesheet_items = fetch_all_list_items(
        token,
        site_id,
        timesheet_list_id
    )

    supplier_items = fetch_all_list_items(
        token,
        site_id,
        supplier_list_id
    )

    site_lookup_items = fetch_all_list_items(
        token,
        site_id,
        site_lookup_list_id
    )

    supplier_map = build_lookup_map(
        supplier_items
    )

    site_map = build_lookup_map(
        site_lookup_items
    )

    print(
        f"✅ Supplier map loaded: "
        f"{len(supplier_map)}"
    )

    print(
        f"✅ Site map loaded: "
        f"{len(site_map)}"
    )

    (
        gap_map,
        gap_summary_rows,
        unmatched_signout_rows
    ) = build_gap_map(
        timesheet_items,
        week_start,
        week_end
    )

    daily_results = []
    roster_raw_rows = []

    for item in roster_items:

        fields = item.get("fields", {})

        position = fields.get("Position", "")

        site_name = get_site_value(
            fields,
            site_map
        )

        supplier_name = get_supplier_value(
            fields,
            supplier_map
        )

        if is_asset_transport_hire(
            position,
            site_name
        ):
            continue

        opms = str(
            fields.get("OPMS", "")
        ).strip()

        roster_date_raw = (
            fields.get("RosterDate")
            or fields.get("Date_x0020_From")
            or fields.get("Date")
        )

        if not opms or not roster_date_raw:
            continue

        roster_date = parse_sharepoint_date(
            roster_date_raw
        )

        if (
            roster_date < week_start
            or roster_date > week_end
        ):
            continue

        work_type = (
            fields.get("WorkType")
            or fields.get("Work Type")
            or ""
        )

        roster_hours = get_roster_hours(
            fields
        )

        shift_type = get_shift_type(
            work_type
        )

        base_row = {
            "Name": fields.get("Title", ""),
            "Position": position,
            "Project": get_project_value(fields),
            "Site": site_name,
            "Supplier": supplier_name,
            "OPMS": opms,
            "RosterDate": roster_date,
            "ShiftType": shift_type,
            "RosterHours": round(roster_hours, 2),
        }

        roster_raw_rows.append(base_row)

        key = (
            opms,
            roster_date
        )

        gap_hours = gap_map.get(
            key,
            0.0
        )

        actual_hours = (
            roster_hours - gap_hours
        )

        if actual_hours < 0:
            actual_hours = 0

        row = {
            **base_row,
            "GapHours": round(gap_hours, 2),
            "ActualHours": round(actual_hours, 2),
        }

        daily_results.append(row)

    print(
        f"✅ Roster raw rows calculated: "
        f"{len(roster_raw_rows)}"
    )

    print(
        f"✅ Daily rows calculated: "
        f"{len(daily_results)}"
    )

    return (
        daily_results,
        roster_raw_rows,
        gap_summary_rows,
        unmatched_signout_rows,
        week_start,
        week_end
    )


# ======================================================
# SUMMARY BUILDERS
# ======================================================

def build_weekly_rows(
    daily_results,
    week_start
):

    people = {}

    for row in daily_results:

        opms = row["OPMS"]

        work_date = row["RosterDate"]

        day_index = (
            work_date - week_start
        ).days

        if day_index < 0 or day_index > 6:
            continue

        if opms not in people:

            people[opms] = {
                "Name": row["Name"],
                "Position": row["Position"],
                "Project": row["Project"],
                "Site": row["Site"],
                "Supplier": row["Supplier"],
                "days": {
                    i: {
                        "DS": 0.0,
                        "NS": 0.0,
                    }
                    for i in range(7)
                },
                "TotalDS": 0.0,
                "TotalNS": 0.0,
                "GrandTotal": 0.0,
            }

        shift = row["ShiftType"]

        people[opms]["days"][day_index][shift] += (
            row["ActualHours"]
        )

    return finalize_people_rows(people)


def build_roster_summary_rows(
    roster_raw_rows,
    week_start
):

    people = {}

    for row in roster_raw_rows:

        opms = row["OPMS"]

        work_date = row["RosterDate"]

        day_index = (
            work_date - week_start
        ).days

        if day_index < 0 or day_index > 6:
            continue

        if opms not in people:

            people[opms] = {
                "Name": row["Name"],
                "Position": row["Position"],
                "Project": row["Project"],
                "Site": row["Site"],
                "Supplier": row["Supplier"],
                "days": {
                    i: {
                        "DS": 0.0,
                        "NS": 0.0,
                    }
                    for i in range(7)
                },
                "TotalDS": 0.0,
                "TotalNS": 0.0,
                "GrandTotal": 0.0,
            }

        shift = row["ShiftType"]

        people[opms]["days"][day_index][shift] += (
            row["RosterHours"]
        )

    return finalize_people_rows(people)


def finalize_people_rows(people):

    final_rows = []

    for person in people.values():

        total_ds = 0.0
        total_ns = 0.0

        for i in range(7):

            ds = round(
                person["days"][i]["DS"],
                2
            )

            ns = round(
                person["days"][i]["NS"],
                2
            )

            person["days"][i]["DS"] = ds
            person["days"][i]["NS"] = ns

            total_ds += ds
            total_ns += ns

        person["TotalDS"] = round(
            total_ds,
            2
        )

        person["TotalNS"] = round(
            total_ns,
            2
        )

        person["GrandTotal"] = round(
            total_ds + total_ns,
            2
        )

        if person["GrandTotal"] <= 0:
            continue

        final_rows.append(person)

    final_rows.sort(
        key=lambda x: x["Name"]
    )

    return final_rows


# ======================================================
# EXCEL HELPERS
# ======================================================

def auto_width_all_sheets(wb):

    for sheet in wb.worksheets:

        for column_cells in sheet.columns:

            max_length = 0

            column_letter = get_column_letter(
                column_cells[0].column
            )

            for cell in column_cells:

                try:
                    if cell.value:
                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )
                except:
                    pass

            sheet.column_dimensions[
                column_letter
            ].width = min(
                max_length + 2,
                35
            )


def write_weekly_style_sheet(
    wb,
    sheet_name,
    rows,
    week_start,
    week_end,
    title_text
):

    ws = wb.create_sheet(sheet_name)

    ws["B2"] = title_text

    ws["B2"].font = Font(
        size=18,
        bold=True
    )

    ws["Q2"] = "Week Ending"

    ws["R2"] = week_end.strftime(
        "%A, %d %B %Y"
    )

    start_row = 6

    ws.cell(start_row, 1, "Name")
    ws.cell(start_row, 2, "OPMS Position")
    ws.cell(start_row, 3, "Client Name")
    ws.cell(start_row, 4, "Supplier")

    days = [
        "Mon",
        "Tue",
        "Wed",
        "Thu",
        "Fri",
        "Sat",
        "Sun",
    ]

    col = 5

    for i, day_name in enumerate(days):

        current_date = (
            week_start + timedelta(days=i)
        )

        ws.merge_cells(
            start_row=start_row,
            start_column=col,
            end_row=start_row,
            end_column=col + 1,
        )

        ws.cell(start_row, col, day_name)

        ws.merge_cells(
            start_row=start_row + 1,
            start_column=col,
            end_row=start_row + 1,
            end_column=col + 1,
        )

        ws.cell(
            start_row + 1,
            col,
            current_date.strftime("%d/%m/%Y")
        )

        ws.cell(
            start_row + 2,
            col,
            "DS"
        )

        ws.cell(
            start_row + 2,
            col + 1,
            "NS"
        )

        col += 2

    total_col = col

    ws.merge_cells(
        start_row=start_row,
        start_column=total_col,
        end_row=start_row + 1,
        end_column=total_col + 1,
    )

    ws.cell(
        start_row,
        total_col,
        "Total"
    )

    ws.cell(
        start_row + 2,
        total_col,
        "DS"
    )

    ws.cell(
        start_row + 2,
        total_col + 1,
        "NS"
    )

    grand_total_col = total_col + 2

    ws.merge_cells(
        start_row=start_row,
        start_column=grand_total_col,
        end_row=start_row + 2,
        end_column=grand_total_col,
    )

    ws.cell(
        start_row,
        grand_total_col,
        "Grand Total"
    )

    row_num = start_row + 3

    for person in rows:

        ws.cell(row_num, 1, person["Name"])

        ws.cell(
            row_num,
            2,
            person["Position"]
        )

        ws.cell(
            row_num,
            3,
            person["Project"]
        )

        ws.cell(
            row_num,
            4,
            person["Supplier"]
        )

        col = 5

        for i in range(7):

            ds = person["days"][i]["DS"]
            ns = person["days"][i]["NS"]

            ws.cell(
                row_num,
                col,
                ds if ds > 0 else ""
            )

            ws.cell(
                row_num,
                col + 1,
                ns if ns > 0 else ""
            )

            col += 2

        ws.cell(
            row_num,
            total_col,
            person["TotalDS"]
        )

        ws.cell(
            row_num,
            total_col + 1,
            person["TotalNS"]
        )

        ws.cell(
            row_num,
            grand_total_col,
            person["GrandTotal"]
        )

        row_num += 1


def write_gap_summary_sheet(
    wb,
    gap_summary_rows,
    unmatched_signout_rows
):

    ws = wb.create_sheet("Gap Summary")

    # ======================================================
    # LEFT TABLE: GAP SUMMARY
    # ======================================================

    gap_headers = [
        "OPMS",
        "Date",
        "TotalGapHours",
    ]

    for col_num, header in enumerate(gap_headers, 1):

        cell = ws.cell(
            1,
            col_num,
            header
        )

        cell.font = Font(bold=True)

    for row_num, row in enumerate(gap_summary_rows, 2):

        ws.cell(
            row_num,
            1,
            row.get("OPMS", "")
        )

        work_date = row.get("Date", "")

        if isinstance(work_date, date):
            work_date = work_date.strftime("%Y/%m/%d")

        ws.cell(
            row_num,
            2,
            work_date
        )

        ws.cell(
            row_num,
            3,
            row.get("TotalGapHours", 0)
        )

    # ======================================================
    # RIGHT TABLE: UNMATCHED SIGN OUT
    # ======================================================

    start_col = 5

    unmatched_headers = [
        "OPMS",
        "SignOutTime",
        "NextSignInTime",
        "HoursBetween",
        "Reason",
    ]

    for index, header in enumerate(unmatched_headers):

        cell = ws.cell(
            1,
            start_col + index,
            header
        )

        cell.font = Font(bold=True)

    for row_num, row in enumerate(unmatched_signout_rows, 2):

        ws.cell(
            row_num,
            start_col,
            row.get("OPMS", "")
        )

        sign_out_time = row.get(
            "SignOutTime",
            ""
        )

        if isinstance(sign_out_time, datetime):
            sign_out_time = sign_out_time.strftime(
                "%Y/%m/%d %H:%M:%S"
            )

        ws.cell(
            row_num,
            start_col + 1,
            sign_out_time
        )

        next_sign_in_time = row.get(
            "NextSignInTime",
            ""
        )

        if isinstance(next_sign_in_time, datetime):
            next_sign_in_time = next_sign_in_time.strftime(
                "%Y/%m/%d %H:%M:%S"
            )

        ws.cell(
            row_num,
            start_col + 2,
            next_sign_in_time
        )

        ws.cell(
            row_num,
            start_col + 3,
            row.get("HoursBetween", "")
        )

        ws.cell(
            row_num,
            start_col + 4,
            row.get("Reason", "")
        )


# ======================================================
# EXCEL EXPORT
# ======================================================

def export_excel(
    weekly_rows,
    roster_summary_rows,
    gap_summary_rows,
    unmatched_signout_rows,
    week_start,
    week_end,
    output_file
):

    wb = Workbook()

    default_ws = wb.active
    wb.remove(default_ws)

    # 1. Weekly Timesheet - 扣完 gap 后
    write_weekly_style_sheet(
        wb,
        "Weekly Timesheet",
        weekly_rows,
        week_start,
        week_end,
        "Weekly Timesheet"
    )

    # 2. Roster Summary - 原始 roster，不扣 gap
    write_weekly_style_sheet(
        wb,
        "Roster Summary",
        roster_summary_rows,
        week_start,
        week_end,
        "Roster Summary"
    )

    # 3. Gap Summary - 左边正常 gap，右边 unmatched sign out
    write_gap_summary_sheet(
        wb,
        gap_summary_rows,
        unmatched_signout_rows
    )

    auto_width_all_sheets(wb)

    format_report_workbook(wb)

    excel_stream = BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)

    print(f"✅ Excel generated in memory: {output_file}")

    return excel_stream.getvalue()


# ======================================================
# MAIN
# ======================================================

def generate_weekly_timesheet():

    (
        daily_results,
        roster_raw_rows,
        gap_summary_rows,
        unmatched_signout_rows,
        week_start,
        week_end
    ) = calculate_daily_results()

    weekly_rows = build_weekly_rows(
        daily_results,
        week_start
    )

    roster_summary_rows = build_roster_summary_rows(
        roster_raw_rows,
        week_start
    )

    output_file = (
        f"Weekly_Timesheet_"
        f"{week_start.strftime('%Y%m%d')}_"
        f"{week_end.strftime('%Y%m%d')}.xlsx"
    )

    export_excel(
        weekly_rows,
        roster_summary_rows,
        gap_summary_rows,
        unmatched_signout_rows,
        week_start,
        week_end,
        output_file
    )

    return output_file, week_start, week_end


if __name__ == "__main__":
    generate_weekly_timesheet()